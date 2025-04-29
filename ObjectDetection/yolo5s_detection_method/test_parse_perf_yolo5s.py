#!/usr/bin/env python3
import re
import sys
import os
from openpyxl import Workbook, load_workbook

# Define the metrics we care about and their regexes:
PATTERNS = {
    "Instructions": r"([\d,]+)\s+instructions\b",
    "Cycles":       r"([\d,]+)\s+cycles\b",
    "IPC":          r"([\d.]+)\s+insn per cycle",
    "Elapsed_s":    r"([\d.]+)\s+seconds\s+time\s+elapsed",
    "User_s":       r"([\d.]+)\s+seconds\s+user\b",
    "Sys_s":        r"([\d.]+)\s+seconds\s+sys\b",
}

def parse_perf_file(path):
    # Read everything at once
    text = open(path, "r").read()
    perf = {}
    for key, pat in PATTERNS.items():
        m = re.search(pat, text, re.IGNORECASE)
        if not m:
            continue
        raw = m.group(1).replace(",", "")
        if key in ("Instructions", "Cycles"):
            perf[key] = int(raw)
        else:
            perf[key] = float(raw)

    # Derive instructions/sec if possible
    if "Instructions" in perf and "Elapsed_s" in perf:
        perf["Insns_per_s"] = perf["Instructions"] / perf["Elapsed_s"]

    return perf

def append_to_excel(out_path, run_number, perf):
    headers = [
        "Run Number", "Instructions", "Cycles", "IPC",
        "Elapsed Time (s)", "Insns per Second",
        "User Time (s)", "Sys Time (s)",
    ]
    row = [
        run_number,
        perf.get("Instructions", "N/A"),
        perf.get("Cycles",       "N/A"),
        perf.get("IPC",          "N/A"),
        perf.get("Elapsed_s",    "N/A"),
        perf.get("Insns_per_s",  "N/A"),
        perf.get("User_s",       "N/A"),
        perf.get("Sys_s",        "N/A"),
    ]

    if os.path.exists(out_path):
        wb = load_workbook(out_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    ws.append(row)
    wb.save(out_path)

def main():
    if len(sys.argv) != 3:
        print(f"usage: {sys.argv[0]} <perf.txt> <run#>")
        sys.exit(1)

    perf_file = sys.argv[1]
    run = int(sys.argv[2])
    perf = parse_perf_file(perf_file)

    OUT = "../assets/results_yolo5s_method/perf_metrics.xlsx"
    append_to_excel(OUT, run, perf)
    print(f"Appended run {run} → {OUT}")

if __name__ == "__main__":
    main()

