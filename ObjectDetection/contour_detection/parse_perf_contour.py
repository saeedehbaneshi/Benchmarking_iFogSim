#!/usr/bin/env python3
import json, re, os, sys
from openpyxl import Workbook, load_workbook

def parse_perf_file(path):
    m = {}
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            # match "<number> <metric>"
            pm = re.match(r"^([\d,.]+)\s+(\S+)", line)
            if pm:
                val = pm.group(1).replace(",", "")
                key = pm.group(2).lower()
                if key=="instructions":
                    m["instructions"] = int(val)
                    ipc = re.search(r"([\d.]+)\s+insn per cycle", line.lower())
                    if ipc:
                        m["ipc"] = float(ipc.group(1))
                elif key=="cycles":
                    m["cycles"] = int(val)
                elif "seconds time elapsed" in line.lower():
                    dt = re.search(r"(\d+\.\d+)", line)
                    if dt:
                        m["elapsed_perf_s"] = float(dt.group(1))
    return m

if __name__=="__main__":
    if len(sys.argv)!=4:
        print("Usage: parse_perf_contour.py <json_metrics> <perf_txt> <run#>")
        sys.exit(1)

    json_path, perf_path, run = sys.argv[1], sys.argv[2], int(sys.argv[3])
    with open(json_path) as f:
        det = json.load(f)
    perf = parse_perf_file(perf_path)

    instr = perf.get("instructions", 0)
    dt_perf = perf.get("elapsed_perf_s", 0)
    fc = det["frame_count"]
    ips = instr/dt_perf if dt_perf>0 else 0
    ipf = instr/fc      if fc>0 else 0

    row = {
        "Run": run,
        "In_Frames": det["frame_count"],
        "In_MB":     det["input_mb"],
        "Out_Frames":det["output_frames"],
        "Out_MB":    det["output_mb"],
        "Elapsed_s": det["elapsed_s"],
        "FPS":       det["fps"],
        "CPU%_per_core": det["cpu_pct_per_core"],
        "Mem_delta_MB": det["mem_delta_mb"],
        "Det_Frames": det["detection_frames"],
        "Total_Dets": det["total_detections"],
        "Det_rate_%":det["detection_rate_pct"],
        "Instr":     perf.get("instructions","N/A"),
        "Cycles":    perf.get("cycles","N/A"),
        "IPC":       perf.get("ipc","N/A"),
        "Instr_per_s": ips,
        "Instr_per_f": ipf,
        "Perf_elapsed_s": perf.get("elapsed_perf_s","N/A")
    }

    OUT = "../assets/results_contour_method/perf_metrics.xlsx"
    if os.path.exists(OUT):
        wb = load_workbook(OUT); ws = wb.active
    else:
        wb = Workbook(); ws = wb.active
        headers = list(row.keys())
        ws.append(headers)

    ws.append(list(row.values()))
    wb.save(OUT)
    print(f"Appended run {run} → {OUT}")

