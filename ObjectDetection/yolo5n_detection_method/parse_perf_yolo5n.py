#!/usr/bin/env python3
import json, re, sys, os
from openpyxl import Workbook, load_workbook

def parse_perf_file(path):
    perf = {}
    with open(path) as f:
        for line in f:
            line = line.strip()
            m = re.match(r"^([\d,.]+)\s+(\S+)", line)
            if not m: continue
            val = m.group(1).replace(",", "")
            key = m.group(2).lower()
            if key=="instructions":
                perf["instructions"] = int(val)
                ipc = re.search(r"([\d.]+)\s+insn per cycle", line.lower())
                if ipc: perf["ipc"] = float(ipc.group(1))
            elif key=="cycles":
                perf["cycles"] = int(val)
            elif "seconds time elapsed" in line.lower():
                t = re.search(r"(\d+\.\d+)", line)
                if t: perf["elapsed_perf_s"] = float(t.group(1))
    return perf

if __name__=="__main__":
    if len(sys.argv)!=4:
        print("usage: parse_perf_yolo5n.py <json> <perf.txt> <run#>")
        sys.exit(1)

    jm, pm, run = sys.argv[1], sys.argv[2], int(sys.argv[3])
    with open(jm) as f: d = json.load(f)
    p = parse_perf_file(pm)

    instr = p.get("instructions",0)
    pt    = p.get("elapsed_perf_s",0)
    fc    = d["frame_count"]
    ips   = instr/pt if pt>0 else 0
    ipf   = instr/fc if fc>0 else 0

    row = {
      "Run Number":            run,
      "Input Frames":          d["frame_count"],
      "Input File Size (MB)":  d["input_mb"],
      "Output Frames":         d["output_frames"],
      "Output File Size (MB)": d["output_mb"],
      "Processing Time (s)":   d["elapsed_s"],
      "FPS":                   d["fps"],
      "CPU Usage per Core (%)":d["cpu_pct_per_core"],
      "Memory Usage (MB)":     d["mem_delta_mb"],
      "Detection Frames":      d["detection_frames"],
      "Total Detections":      d["total_detections"],
      "Detection Rate (%)":    d["detection_rate_pct"],
      "Instructions":          p.get("instructions","N/A"),
      "Cycles":                p.get("cycles","N/A"),
      "Insn per Cycle":        p.get("ipc","N/A"),
      "Instructions per Second": ips,
      "Instructions per Frame":  ipf,
      "Perf Time Elapsed (s)":  p.get("elapsed_perf_s","N/A")
    }

    OUT = "../assets/results_yolo5n_method/perf_metrics.xlsx"
    if os.path.exists(OUT):
        wb = load_workbook(OUT); ws = wb.active
    else:
        wb = Workbook(); ws = wb.active
        ws.append(list(row.keys()))
    ws.append(list(row.values()))
    wb.save(OUT)
    print(f"Appended run {run} → {OUT}")

