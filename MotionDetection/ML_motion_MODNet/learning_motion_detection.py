import cv2
import torch
import numpy as np
import time
import os
import sys
import subprocess
from openpyxl import Workbook, load_workbook
from torchvision import transforms

# Add MODNet's `src` directory to Python path
sys.path.append(os.path.abspath('./MODNet/src'))

from models.modnet import MODNet  # Import MODNet architecture from the repository

# Paths
video_path = "Sample_video.mp4"
output_video_path = "deep_motion_mask.avi"
excel_output_path = "perf_metrics_deep_learning.xlsx"
backbone_path = "./MODNet/pretrained/mobilenetv2_human_seg.ckpt"
perf_output_path = "perf_output.txt"

# Load the device (GPU or CPU)
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# Check if backbone model exists
if not os.path.exists(backbone_path):
    print(f"Error: Backbone model not found at {backbone_path}")
    exit()

print(f"Loading backbone model from: {backbone_path}")
backbone = torch.load(backbone_path, map_location=device)

# Load MODNet with manually assigned backbone
model = MODNet(backbone_pretrained=False).to(device)  # Disable default backbone loading
model.load_state_dict(backbone, strict=False)  # Manually load backbone
model.eval()

# Load the video file
cap = cv2.VideoCapture(video_path)
if not cap.isOpened():
    print("Error: Unable to open video file.")
    exit()

# Get input video properties
frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
fps = int(cap.get(cv2.CAP_PROP_FPS))
frame_count_total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
input_file_size = os.path.getsize(video_path) / (1024 ** 2)  # File size in MB

# Create a VideoWriter for saving the motion mask
fourcc = cv2.VideoWriter_fourcc(*"XVID")
out = cv2.VideoWriter(output_video_path, fourcc, fps, (frame_width, frame_height), isColor=False)

# Initialize metrics
frame_counter = 0
motion_frames = 0
total_motion_area = 0

# Transform for model input
transform = transforms.Compose([
    transforms.ToTensor(),
    transforms.Resize((512, 512)),  # Adjust the input size to match MODNet requirements
    transforms.Normalize(mean=[0.5, 0.5, 0.5], std=[0.5, 0.5, 0.5])  # Normalize for RGB
])

# Measure processing time
start_time = time.time()

# Run perf command
perf_command = f"perf stat -e instructions -- python3 -c 'import time; time.sleep(0.1)'"
with open(perf_output_path, "w") as perf_file:
    subprocess.run(perf_command, shell=True, stderr=perf_file)

# Extract instructions count
instructions = 0
with open(perf_output_path, "r") as perf_file:
    for line in perf_file:
        if "instructions" in line:
            instructions = int(line.split()[0].replace(",", ""))
            break

while cap.isOpened():
    ret, frame = cap.read()
    if not ret:
        break

    # Preprocess frame for model input
    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)  # Convert to RGB
    input_tensor = transform(frame_rgb).unsqueeze(0).to(device)

    # Perform motion detection using MODNet
    with torch.no_grad():
        _, _, motion_mask = model(input_tensor, inference=True)

    # Convert model output to binary motion mask
    motion_mask = (motion_mask.squeeze().cpu().numpy() > 0.5).astype(np.uint8) * 255
    motion_mask = cv2.resize(motion_mask, (frame_width, frame_height))  # Resize to original frame size

    # Calculate motion statistics
    motion_area = (motion_mask == 255).sum()
    if motion_area > 0:
        motion_frames += 1
        total_motion_area += motion_area

    # Write the motion mask to the output file
    out.write(motion_mask)

    frame_counter += 1

# Calculate processing time
elapsed_time = time.time() - start_time
calculated_fps = frame_counter / elapsed_time

# Release resources
cap.release()
out.release()

# Save metrics to Excel
if os.path.exists(excel_output_path):
    wb = load_workbook(excel_output_path)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Perf Metrics"
    # Write headers if this is a new file
    ws.append([
        "Run Number",
        "Input Frames",
        "Input File Size (MB)",
        "Output Frames",
        "Processing Time (s)",
        "FPS",
        "Motion Frames",
        "Motion Area (Total Pixels)",
        "Instructions"
    ])

# Determine the run number
run_number = ws.max_row

# Append data
ws.append([
    run_number,                      # Run number
    frame_count_total,               # Input frames
    input_file_size,                 # Input file size
    frame_counter,                   # Output frames
    elapsed_time,                    # Processing time
    calculated_fps,                  # FPS
    motion_frames,                   # Motion frames
    total_motion_area,               # Total motion area (pixels)
    instructions                     # Number of instructions
])

# Save the workbook
wb.save(excel_output_path)
print(f"Performance metrics appended to Excel file: {excel_output_path}")
print(f"Processed {frame_counter} frames in {elapsed_time:.2f} seconds (FPS: {calculated_fps:.2f})")
print(f"Motion mask video saved at: {output_video_path}")

