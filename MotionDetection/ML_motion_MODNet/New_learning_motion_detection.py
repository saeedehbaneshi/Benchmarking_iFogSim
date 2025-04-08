#!/usr/bin/env python3
"""
ML-Based Motion Detection using MODNet with Improved Post-Processing

This script uses MODNet (with a MobileNetV2 backbone) to generate a motion mask
from an input video (e.g., containing many moving cars and motorcycles).
It then applies additional post-processing (median blur and morphological closing)
to produce a cleaner, less noisy binary mask.
Finally, the processed mask is saved as an output video and some performance metrics are logged.

Before running:
- Ensure the MODNet repository is in ./MODNet and that its `src` directory is accessible.
- Verify that the backbone checkpoint file exists at ./MODNet/pretrained/mobilenetv2_human_seg.ckpt.
- Place your input video (with moving vehicles) in the same folder or update video_path accordingly.
"""

import cv2
import torch
import numpy as np
import time
import os
import sys
import subprocess
from openpyxl import Workbook, load_workbook
from torchvision import transforms

# ---------------------------
# 1. Setup and Model Loading
# ---------------------------

# Add MODNet's 'src' directory to Python path
sys.path.append(os.path.abspath('./MODNet/src'))
from models.modnet import MODNet  # Import MODNet architecture

# File paths (update these paths if necessary)
video_path = "Sample_video.mp4"               # Input video path
output_video_path = "new_deep_motion_mask.avi"      # Output mask video
excel_output_path = "perf_metrics_deep_learning.xlsx"
backbone_path = "./MODNet/pretrained/mobilenetv2_human_seg.ckpt"
perf_output_path = "perf_output.txt"

# Set detection threshold (tune this value for clean binary mask)
THRESHOLD = 0.5

# Load the device (GPU if available, else CPU)
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# Check if the backbone model file exists
if not os.path.exists(backbone_path):
    print(f"Error: Backbone model not found at {backbone_path}")
    exit()

print(f"Loading backbone model from: {backbone_path}")
backbone = torch.load(backbone_path, map_location=device)

# Load MODNet with the manually assigned backbone (disable default loading)
model = MODNet(backbone_pretrained=False).to(device)
model.load_state_dict(backbone, strict=False)
model.eval()

# -------------------------
# 2. Video I/O Setup
# -------------------------

cap = cv2.VideoCapture(video_path)
if not cap.isOpened():
    print("Error: Unable to open video file.")
    exit()

frame_width  = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
fps = int(cap.get(cv2.CAP_PROP_FPS))
frame_count_total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
input_file_size = os.path.getsize(video_path) / (1024 ** 2)  # in MB

# Define VideoWriter (we output a single-channel video, hence isColor=False)
fourcc = cv2.VideoWriter_fourcc(*"XVID")
out = cv2.VideoWriter(output_video_path, fourcc, fps, (frame_width, frame_height), isColor=False)

# -------------------------------
# 3. Define Preprocessing Transform
# -------------------------------
# MODNet was trained with images resized to 512x512 and normalized with mean=0.5, std=0.5.
transform = transforms.Compose([
    transforms.ToTensor(),
    transforms.Resize((512, 512)),
    transforms.Normalize(mean=[0.5, 0.5, 0.5], std=[0.5, 0.5, 0.5])
])

# -------------------------------
# 4. Performance Metrics Setup
# -------------------------------
frame_counter = 0
motion_frames = 0
total_motion_area = 0

start_time = time.time()

# Run a short perf command and log (optional)
perf_command = f"perf stat -e instructions -- python3 -c 'import time; time.sleep(0.1)'"
with open(perf_output_path, "w") as perf_file:
    subprocess.run(perf_command, shell=True, stderr=perf_file)

instructions = 0
with open(perf_output_path, "r") as perf_file:
    for line in perf_file:
        if "instructions" in line:
            instructions = int(line.split()[0].replace(",", ""))
            break

# -------------------------------
# 5. Process Each Frame
# -------------------------------
while cap.isOpened():
    ret, frame = cap.read()
    if not ret:
        break

    # Convert BGR frame to RGB for MODNet input
    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    input_tensor = transform(frame_rgb).unsqueeze(0).to(device)

    # Inference: Get the motion mask from MODNet.
    # MODNet returns three outputs; the third one is our motion mask.
    with torch.no_grad():
        _, _, motion_mask = model(input_tensor, inference=True)

    # Convert output tensor to numpy array and apply threshold.
    # The output is resized to (512,512); resize it back to original frame size.
    motion_mask = motion_mask.squeeze().cpu().numpy()
    binary_mask = (motion_mask > THRESHOLD).astype(np.uint8) * 255
    binary_mask = cv2.resize(binary_mask, (frame_width, frame_height))

    # -------------------------------
    # 6. Post-Processing to Improve Mask Quality
    # -------------------------------
    # Apply a median blur to reduce salt-and-pepper noise.
    binary_mask = cv2.medianBlur(binary_mask, 5)
    # Apply morphological closing to fill small holes (kernel size can be tuned)
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
    binary_mask = cv2.morphologyEx(binary_mask, cv2.MORPH_CLOSE, kernel)

    # Calculate motion area (number of pixels with value 255)
    motion_area = np.sum(binary_mask == 255)
    if motion_area > 0:
        motion_frames += 1
        total_motion_area += motion_area

    # Write the processed binary mask to output video
    out.write(binary_mask)
    frame_counter += 1

    # (Optional) Display current mask for debugging
    cv2.imshow("Processed Motion Mask", binary_mask)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# -------------------------------
# 7. Post-Processing and Saving Metrics
# -------------------------------
elapsed_time = time.time() - start_time
calculated_fps = frame_counter / elapsed_time

cap.release()
out.release()
cv2.destroyAllWindows()

# Save performance metrics to Excel
if os.path.exists(excel_output_path):
    wb = load_workbook(excel_output_path)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Perf Metrics"
    ws.append([
        "Run Number", "Input Frames", "Input File Size (MB)", "Output Frames",
        "Processing Time (s)", "FPS", "Motion Frames", "Motion Area (Total Pixels)",
        "Instructions"
    ])

run_number = ws.max_row
ws.append([
    run_number,
    frame_count_total,
    input_file_size,
    frame_counter,
    elapsed_time,
    calculated_fps,
    motion_frames,
    total_motion_area,
    instructions
])
wb.save(excel_output_path)

print(f"Performance metrics appended to Excel file: {excel_output_path}")
print(f"Processed {frame_counter} frames in {elapsed_time:.2f} seconds (FPS: {calculated_fps:.2f})")
print(f"Motion mask video saved at: {output_video_path}")

