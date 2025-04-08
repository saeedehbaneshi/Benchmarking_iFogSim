import cv2
import time
import psutil
import os
import subprocess
from openpyxl import Workbook, load_workbook

# Paths
video_path = "Sample_video.mp4"
output_video_path = "motion_mask_output.avi"
excel_output_path = "perf_metrics.xlsx"
perf_output_path = "perf_output.txt"

# Load the video file
cap = cv2.VideoCapture(video_path)

if not cap.isOpened():
    print("Error: Unable to open video file.")
    exit()

# Get input video properties
frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
fps = int(cap.get(cv2.CAP_PROP_FPS))
frame_count_total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
input_file_size = os.path.getsize(video_path) / (1024 ** 2)  # File size in MB

# Create a VideoWriter for saving the motion mask
fourcc = cv2.VideoWriter_fourcc(*"XVID")
out = cv2.VideoWriter(output_video_path, fourcc, fps, (frame_width, frame_height), isColor=False)

# Create Background Subtractor
bg_subtractor = cv2.createBackgroundSubtractorMOG2(history=500, varThreshold=50, detectShadows=True)

# Measure processing time and system performance
frame_count = 0
start_time = time.time()
process = psutil.Process()

cpu_start = process.cpu_percent(interval=None)
memory_start = process.memory_info().rss  # Memory usage in bytes

# Motion detection statistics
motion_frames = 0
total_motion_area = 0

# Run `perf` to measure instructions, cycles, etc.
perf_command = f"perf stat -e instructions,cycles,cache-misses,branch-misses -- python3 -c 'import time; time.sleep(0.001)'"
with open(perf_output_path, "w") as perf_file:
    subprocess.run(perf_command, shell=True, stderr=perf_file)

while cap.isOpened():
    ret, frame = cap.read()
    if not ret:
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    motion_mask = bg_subtractor.apply(gray)
    out.write(motion_mask)

    # Calculate motion statistics
    motion_area = (motion_mask == 255).sum()  # Count white pixels
    if motion_area > 0:
        motion_frames += 1
        total_motion_area += motion_area

    frame_count += 1
    #print(f"Writing frame {frame_count} to output.")

# Measure elapsed time and system metrics
elapsed_time = time.time() - start_time
calculated_fps = frame_count / elapsed_time

cpu_end = process.cpu_percent(interval=None)
memory_end = process.memory_info().rss  # Memory usage in bytes

# Output video properties
output_file_size = os.path.getsize(output_video_path) / (1024 ** 2)  # File size in MB
output_cap = cv2.VideoCapture(output_video_path)
output_frame_count = int(output_cap.get(cv2.CAP_PROP_FRAME_COUNT))
output_cap.release()

# Calculate motion percentage
motion_percentage = (motion_frames / frame_count) * 100

# Extract perf metrics
perf_metrics = {}
with open(perf_output_path, "r") as perf_file:
    for line in perf_file:
        if "instructions" in line:
            perf_metrics["instructions"] = int(line.split()[0].replace(",", ""))
        elif "cycles" in line:
            perf_metrics["cycles"] = int(line.split()[0].replace(",", ""))
        elif "cache-misses" in line:
            perf_metrics["cache_misses"] = int(line.split()[0].replace(",", ""))
        elif "branch-misses" in line:
            perf_metrics["branch_misses"] = int(line.split()[0].replace(",", ""))
        elif "seconds time elapsed" in line:
            perf_metrics["perf_time_elapsed"] = float(line.split()[0])

# Release resources
cap.release()
out.release()

print(f"Processed {frame_count} frames in {elapsed_time:.2f} seconds (FPS: {calculated_fps:.2f})")
print(f"Motion mask video saved at: {output_video_path}")

# Save metrics to Excel
if os.path.exists(excel_output_path):
    # If file exists, load it
    wb = load_workbook(excel_output_path)
    ws = wb.active
else:
    # If file doesn't exist, create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Perf Metrics"
    # Write headers if this is a new file
    ws.append([
        "Run Number",
        "Input Frames",
        "Input File Size (MB)",
        "Output Frames",
        "Output File Size (MB)",
        "Processing Time (s)",
        "FPS",
        "CPU Usage (%)",
        "Memory Usage (MB)",
        "Motion Frames",
        "Motion Area (Total Pixels)",
        "Motion Percentage (%)",
        "Instructions",
        "Cycles",
        "Cache Misses",
        "Branch Misses",
        "Perf Time Elapsed (s)"
    ])

# Determine the run number
run_number = ws.max_row  # Existing rows determine the run number

# Append data
ws.append([
    run_number,                      # Run number
    frame_count_total,               # Input frames
    input_file_size,                 # Input file size
    output_frame_count,              # Output frames
    output_file_size,                # Output file size
    elapsed_time,                    # Processing time
    calculated_fps,                  # FPS
    cpu_end - cpu_start,             # CPU usage
    (memory_end - memory_start) / (1024 ** 2),  # Memory usage in MB
    motion_frames,                   # Motion frames
    total_motion_area,               # Total motion area (pixels)
    motion_percentage,               # Motion percentage
    perf_metrics.get("instructions", "N/A"),  # Instructions
    perf_metrics.get("cycles", "N/A"),        # Cycles
    perf_metrics.get("cache_misses", "N/A"),  # Cache misses
    perf_metrics.get("branch_misses", "N/A"), # Branch misses
    perf_metrics.get("perf_time_elapsed", "N/A")  # Perf time elapsed
])

# Save the workbook
wb.save(excel_output_path)
print(f"Performance metrics appended to Excel file: {excel_output_path}")

