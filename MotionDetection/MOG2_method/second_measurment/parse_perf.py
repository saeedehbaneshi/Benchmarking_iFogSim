import json
import re
import os
import sys
from openpyxl import load_workbook, Workbook

def parse_perf(perf_file):
    perf_metrics = {}
    with open(perf_file, "r") as pf:
        for line in pf:
            line = line.strip()
            if not line:
                continue
            match = re.match(r"^\s*([\d,.]+)\s+(\S+)", line)
            if match:
                val_str = match.group(1).replace(",", "")
                metric = match.group(2).lower()
                if metric == "instructions":
                    perf_metrics["instructions"] = int(val_str)
                    ipc_match = re.search(r"([\d.]+)\s+insn per cycle", line.lower())
                    if ipc_match:
                        perf_metrics["insn_per_cycle"] = float(ipc_match.group(1))
                elif metric == "cycles":
                    perf_metrics["cycles"] = int(val_str)
                elif "seconds time elapsed" in line:
                    perf_time_match = re.search(r"(\d+\.\d+)", line)
                    if perf_time_match:
                        perf_metrics["perf_time_elapsed"] = float(perf_time_match.group(1))
    return perf_metrics

def main():
    if len(sys.argv) != 4:
        print("Usage: python3 parse_perf.py <json_file> <perf_file> <run_number>")
        sys.exit(1)
    
    json_file = sys.argv[1]
    perf_file = sys.argv[2]
    run_number = int(sys.argv[3])
    
    with open(json_file, "r") as f:
        motion_detection_metrics = json.load(f)
    
    perf_metrics = parse_perf(perf_file)
    
    # Calculate derived metrics
    instructions = perf_metrics.get("instructions", 0)
    perf_time = perf_metrics.get("perf_time_elapsed", 0)
    frame_count = motion_detection_metrics["frame_count"]
    instructions_per_second = instructions / perf_time if perf_time > 0 else 0
    instructions_per_frame = instructions / frame_count if frame_count > 0 else 0
    
    # Combine all metrics
    all_metrics = {
        **motion_detection_metrics,
        **perf_metrics,
        "instructions_per_second": instructions_per_second,
        "instructions_per_frame": instructions_per_frame
    }
    
    # Write to Excel
    EXCEL_OUTPUT_PATH = "perf_metrics.xlsx"
    if os.path.exists(EXCEL_OUTPUT_PATH):
        wb = load_workbook(EXCEL_OUTPUT_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Perf Metrics"
    
    # Define expected headers
    expected_headers = [
        "Run Number",
        "Input Frames",
        "Input File Size (MB)",
        "Output Frames",
        "Output File Size (MB)",
        "Processing Time (s)",
        "FPS",
        "CPU Usage per Core (%)",
        "Memory Usage (MB)",
        "Motion Frames",
        "Total Motion Area",
        "Motion Percentage (%)",
        "Instructions",
        "Cycles",
        "Insn per Cycle",
        "Instructions per Second",
        "Instructions per Frame",
        "Perf Time Elapsed (s)"
    ]
    
    # Check if headers exist in the first row
    if ws.max_row == 0 or [cell.value for cell in ws[1]] != expected_headers:
        if ws.max_row > 0:
            # Clear existing content if headers are incorrect
            ws.delete_rows(1, ws.max_row)
        ws.append(expected_headers)
        print("Headers written to Excel sheet")
    
    # Append the metrics row
    row_data = [
        run_number,
        all_metrics["frame_count_total"],
        all_metrics["input_file_size"],
        all_metrics["output_frame_count"],
        all_metrics["output_file_size"],
        all_metrics["elapsed_time"],
        all_metrics["calculated_fps"],
        all_metrics["cpu_usage_per_core"],
        all_metrics["mem_usage_mb"],
        all_metrics["motion_frames"],
        all_metrics["total_motion_area"],
        all_metrics["motion_percentage"],
        all_metrics.get("instructions", "N/A"),
        all_metrics.get("cycles", "N/A"),
        all_metrics.get("insn_per_cycle", "N/A"),
        all_metrics["instructions_per_second"],
        all_metrics["instructions_per_frame"],
        all_metrics.get("perf_time_elapsed", "N/A")
    ]
    ws.append(row_data)
    
    wb.save(EXCEL_OUTPUT_PATH)
    print(f"Metrics for run {run_number} appended to {EXCEL_OUTPUT_PATH}")

if __name__ == "__main__":
    main()
