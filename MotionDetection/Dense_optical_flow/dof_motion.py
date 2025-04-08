import cv2
import time
import numpy as np
import psutil
import os
from openpyxl import Workbook, load_workbook

# Paths
video_path = "Sample_video.mp4"
output_video_path = "optical_flow_motion_mask.avi"
excel_output_path = "perf_metrics_optical_flow.xlsx"

# Load the video file
cap = cv2.VideoCapture(video_path)
if not cap.isOpened():
    print("Error: Unable to open video file.")
    exit()

# Get input video properties
frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
fps = int(cap.get(cv2.CAP_PROP_FPS))
frame_count_total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
input_file_size = os.path.getsize(video_path) / (1024 ** 2)  # File size in MB

# Create a VideoWriter for saving the motion mask
fourcc = cv2.VideoWriter_fourcc(*"XVID")
out = cv2.VideoWriter(output_video_path, fourcc, fps, (frame_width, frame_height), isColor=False)

# Initialize variables
prev_gray = None
frame_counter = 0
motion_frames = 0
total_motion_area = 0

# Measure processing time and system performance
start_time = time.time()
process = psutil.Process()

cpu_start = process.cpu_percent(interval=None)
memory_start = process.memory_info().rss  # Memory usage in bytes

while cap.isOpened():
    ret, frame = cap.read()
    if not ret:
        break

    # Convert the current frame to grayscale
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    if prev_gray is not None:
        # Calculate dense optical flow using the Farneback method
        flow = cv2.calcOpticalFlowFarneback(
            prev_gray, gray, None,
            pyr_scale=0.5, levels=3, winsize=15,
            iterations=3, poly_n=5, poly_sigma=1.2, flags=0
        )

        # Compute the magnitude and angle of the flow
        magnitude, angle = cv2.cartToPolar(flow[..., 0], flow[..., 1])

        # Normalize the magnitude to [0, 255] for visualization
        motion_mask = cv2.normalize(magnitude, None, 0, 255, cv2.NORM_MINMAX)
        motion_mask = motion_mask.astype(np.uint8)

        # Threshold the motion mask to highlight significant motion
        _, motion_mask = cv2.threshold(motion_mask, 25, 255, cv2.THRESH_BINARY)

        # Calculate motion statistics
        motion_area = (motion_mask == 255).sum()
        if motion_area > 0:
            motion_frames += 1
            total_motion_area += motion_area

        # Write the motion mask to the output file
        out.write(motion_mask)

    # Update the previous frame
    prev_gray = gray
    frame_counter += 1

# Measure elapsed time and system metrics
elapsed_time = time.time() - start_time
calculated_fps = frame_counter / elapsed_time

cpu_end = process.cpu_percent(interval=None)
memory_end = process.memory_info().rss  # Memory usage in bytes

# Output video properties
out.release()
output_file_size = os.path.getsize(output_video_path) / (1024 ** 2)  # File size in MB
output_cap = cv2.VideoCapture(output_video_path)
output_frame_count = int(output_cap.get(cv2.CAP_PROP_FRAME_COUNT))
output_cap.release()

# Calculate motion percentage
motion_percentage = (motion_frames / frame_counter) * 100

print(f"Processed {frame_counter} frames in {elapsed_time:.2f} seconds (FPS: {calculated_fps:.2f})")
print(f"Motion mask video saved at: {output_video_path}")

# Save metrics to Excel
if os.path.exists(excel_output_path):
    # If file exists, load it
    wb = load_workbook(excel_output_path)
    ws = wb.active
else:
    # If file doesn't exist, create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Perf Metrics"
    # Write headers if this is a new file
    ws.append([
        "Run Number",
        "Input Frames",
        "Input File Size (MB)",
        "Output Frames",
        "Output File Size (MB)",
        "Processing Time (s)",
        "FPS",
        "CPU Usage (%)",
        "Memory Usage (MB)",
        "Motion Frames",
        "Motion Area (Total Pixels)",
        "Motion Percentage (%)"
    ])

# Determine the run number
run_number = ws.max_row  # Existing rows determine the run number

# Append data
ws.append([
    run_number,                      # Run number
    frame_count_total,               # Input frames
    input_file_size,                 # Input file size
    output_frame_count,              # Output frames
    output_file_size,                # Output file size
    elapsed_time,                    # Processing time
    calculated_fps,                  # FPS
    cpu_end - cpu_start,             # CPU usage
    (memory_end - memory_start) / (1024 ** 2),  # Memory usage in MB
    motion_frames,                   # Motion frames
    total_motion_area,               # Total motion area (pixels)
    motion_percentage                # Motion percentage
])

# Save the workbook
wb.save(excel_output_path)
print(f"Performance metrics appended to Excel file: {excel_output_path}")

